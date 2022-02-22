import subprocess, datetime, gzip, os, pickle, glob, openpyxl, shutil, psutil
import pandas as pd
import numpy as np
from joblib import Parallel, delayed
from pathlib import Path
from Bio.SeqIO.FastaIO import SimpleFastaParser
from io import StringIO
from tqdm import tqdm
from openpyxl.utils.dataframe import dataframe_to_rows
from functools import reduce

## denoising function to denoise all sequences the input fasta with a given alpha and minsize
def denoise(project = None, comp_lvl = None, cores = None, alpha = None, minsize = None):
    """Function to apply denoisind to a given gzipped file. Outputs a fasta file with all
    centroid sequences."""

    ## define the name for the output fasta
    ## create an output path to write to
    sample_name_out_1 = 'ESVs_with_chimeras.fasta.gz'
    output_path = Path(project).joinpath('8_denoising', 'data', sample_name_out_1)

    ## give user output
    print('{}: Starting denoising. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))

    ## reduce cores to 75% of available ressources to prevent overheating while clustering / denoising:
    if cores > int(psutil.cpu_count() * 0.75):
        cores = int(psutil.cpu_count() * 0.75)

    ## run vsearch --cluster_unoise to cluster OTUs
    ## use --log because for some reason no info is written to stderr with this command
    ## write stdout to uncompressed output at runtime
    with open(output_path.with_suffix(''), 'w') as output:
        f = subprocess.run(['vsearch',
                            '--cluster_unoise', Path(project).joinpath('6_dereplication_pooling', 'data', 'pooling', 'pooled_sequences_dereplicated.fasta.gz'),
                            '--unoise_alpha', str(alpha),
                            '--minsize', str(minsize),
                            '--sizein', '--sizeout',
                            '--centroids', '-', '--fasta_width', str(0), '--quiet',
                            '--log', Path(project).joinpath('8_denoising', 'temp', 'denoising_log.txt'),
                            '--threads', str(cores)], stdout = output, stderr = subprocess.DEVNULL)

    ## compress the output, remove uncompressed output
    with open(output_path.with_suffix(''), 'rb') as in_stream, gzip.open(output_path, 'wb', comp_lvl) as out_stream:
            shutil.copyfileobj(in_stream, out_stream)
    os.remove(output_path.with_suffix(''))

    ## collect processed and passed reads from the log file
    with open(Path(project).joinpath('8_denoising', 'temp', 'denoising_log.txt')) as log_file:
        content = log_file.read().split('\n')
        seqs, esvs = content[3].split(' ')[3], content[17].split(' ')[1]
        version = content[0].split(',')[0]
        finished = '{}'.format(datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    print('{}: Denoised unique {} sequences into {} ESVs.'.format(datetime.datetime.now().strftime("%H:%M:%S"), seqs, esvs))
    print('{}: Starting chimera removal from the OTUs. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))

    ## run vsearch --uchime_denovo to remove chimeric sequences from the OTUs
    f = subprocess.run(['vsearch',
                        '--uchime_denovo', Path(project).joinpath('8_denoising', 'data', sample_name_out_1),
                        '--relabel', 'ESV_',
                        '--nonchimeras', Path(project).joinpath('8_denoising', '{}_ESVs.fasta'.format(Path(project).stem)),
                        '-fasta_width', str(0), '--quiet'])

    ## collect processed and passed reads from the output fasta, since it is not reported in the log
    f = list(SimpleFastaParser(open(Path(project).joinpath('8_denoising', '{}_ESVs.fasta'.format(Path(project).stem)))))
    print('{}: {} chimeras removed from {} ESV sequences.'.format(datetime.datetime.now().strftime("%H:%M:%S"), int(esvs) - len(f), esvs))
    print('{}: ESVs saved to {}.'.format(datetime.datetime.now().strftime("%H:%M:%S"), Path(project).joinpath('7_otu_clustering', '{}_OTUs.fasta'.format(Path(project).stem))))

## remapping function to remap the individual reads to the ESVs via vsearch
def remapping_esv(file, project = None):
    """Function to remap the sequences of a dereplicated file against the ESV list
    as database."""

    ## extract the sample name from the file name for the otu table
    sample_name_out = '{}'.format(Path(file).with_suffix('').with_suffix('').name).replace('_PE_trimmed_filtered_dereplicated', '')

    ## run vsearch --search_exact to remap the individual files vs the generated
    ## ESV fasta, capture log and directly pickle the output as dataframe for read table generation
    f = subprocess.run(['vsearch',
                        '--search_exact', Path(file),
                        '--db', Path(project).joinpath('8_denoising', '{}_ESVs.fasta'.format(Path(project).stem)),
                        '--output_no_hits',
                        '--maxhits', '1',
                        '--otutabout', '-', '--quiet', '--threads', str(1),
                        '--log', Path(project).joinpath('8_denoising', 'temp', '{}_mapping_log.txt'.format(sample_name_out))], capture_output = True)

    ## directly parse the output to a pandas dataframe
    esv_tab = pd.read_csv(StringIO(f.stdout.decode('ascii', errors = 'ignore')), sep = '\t')

    ## handle empty outputs correctly
    if not esv_tab.empty:
        esv_tab = esv_tab.set_axis(['ID', sample_name_out], axis = 1, inplace = False)
    else:
        esv_tab[sample_name_out] = ""

    ## collect number of esvs from the output, pickle to logs
    with open(Path(project).joinpath('8_denoising', 'temp', '{}_mapping_log.txt'.format(sample_name_out))) as log_file:
        content = log_file.read().split('\n')
        esvs, exact_matches = content[3].split(' ')[3], len(esv_tab)
        version = content[0].split(',')[0]
        finished = '{}'.format(datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    ## give user output
    print('{}: {}: {} exact matches found ({} reads).'.format(datetime.datetime.now().strftime("%H:%M:%S"), sample_name_out, exact_matches, esv_tab[sample_name_out].sum()))

    ## pickle log data first for log generation
    with open(Path(project).joinpath('8_denoising', 'temp', '{}_log.pkl'.format(sample_name_out)), 'wb') as log:
        pickle.dump([sample_name_out, finished, version, exact_matches, esv_tab[sample_name_out].sum()], log)

    ## pickle otu tab dataframes for otu table generation
    with open(Path(project).joinpath('8_denoising', 'temp', '{}_esv_tab.pkl'.format(sample_name_out)), 'wb') as log:
        pickle.dump(esv_tab, log)

## main function for the denoising script
def main(project = Path.cwd()):
    """Main function of the script. Default values can be changed via the input file.
    Will denoise the dataset, perform chimera removal, remap the individual files and
    generate an ESV table."""

    ## create temporal output folder
    try:
        os.mkdir(Path(project).joinpath('8_denoising', 'temp'))
    except FileExistsError:
        pass

    ## collect variables from the settings file
    gen_settings = pd.read_excel(Path(project).joinpath('Settings.xlsx'), sheet_name = '0_general_settings')
    cores, comp_lvl = gen_settings['cores to use'].item(), gen_settings['compression level'].item()

    settings = pd.read_excel(Path(project).joinpath('Settings.xlsx'), sheet_name = '8_denoising')
    alpha, minsize, to_excel = settings['alpha'].item(), settings['minsize'].item(), settings['to excel'].item()

    ## denoise the dataset
    denoise(project = project, comp_lvl = comp_lvl, cores = cores, alpha = alpha, minsize = minsize)

    ## gather files for remapping of ESVs
    input = glob.glob(str(Path(project).joinpath('6_dereplication_pooling', 'data', 'dereplication', '*.fasta.gz')))

    print('{}: Starting to remap {} input files.'.format(datetime.datetime.now().strftime("%H:%M:%S"), len(input)))

    ## run remapping parallelized to speed up the process
    Parallel(n_jobs = cores)(delayed(remapping_esv)(file, project = project) for file in input)

    ## write log for the denoising from pkl logs
    summary_logs = glob.glob(str(Path(project).joinpath('8_denoising', 'temp', '*_log.pkl')))
    summary = [pickle.load(open(line, 'rb')) for line in summary_logs]

    log_df = pd.DataFrame(summary, columns = ['File', 'finished at', 'program version', 'exact matches', 'reads matched'])
    log_df = log_df.sort_values(by = 'File')
    log_df.to_excel(Path(project).joinpath('8_denoising', 'Logfile_8_denoising.xlsx'),
                    index = False,
                    sheet_name = '8_denoising')

    ## add log to the project report
    wb = openpyxl.load_workbook(Path(project).joinpath('Project_report.xlsx'))
    writer = pd.ExcelWriter(Path(project).joinpath('Project_report.xlsx'), engine = 'openpyxl')
    writer.book = wb

    ## write the output
    log_df.to_excel(writer, sheet_name = '8_denoising', index = False)
    wb.save(Path(project).joinpath('Project_report.xlsx'))
    writer.close()

    ## generate OTU table, first extract all OTUs and sequences from fasta file
    esv_list = list(SimpleFastaParser(open(Path(project).joinpath('8_denoising', '{}_ESVs.fasta'.format(Path(project).stem)))))
    esv_table = pd.DataFrame(esv_list, columns = ['ID', 'Seq'])
    seq_col = esv_table.pop('Seq')

    ## extract individual ESV tabs from the clustering output, rename columns correctly, merge individual tabs
    esv_tabs = glob.glob(str(Path(project).joinpath('8_denoising', 'temp', '*_esv_tab.pkl')))
    esv_tabs = [pickle.load(open(tab_file, 'rb')) for tab_file in esv_tabs]
    esv_tabs = [tab.rename(columns = {tab.columns[0] : 'ID'}) for tab in esv_tabs]
    esv_tabs = [pd.merge(esv_table, tab, on = 'ID', how = 'outer').set_index('ID') for tab in tqdm(esv_tabs, desc = 'Generating ESV table')]

    ## collapse all individual dataframes into the ESV table, replace nan values with 0
    esv_table = pd.concat(esv_tabs, axis = 1)
    esv_table = esv_table.reset_index(level = 0).fillna(0)
    esv_table = pd.concat([esv_table[['ID']], esv_table[esv_table.columns.difference(['ID'])].sort_index(axis = 1)], ignore_index = False, axis = 1)

    ## move sequences to the end of the dataframe
    esv_table.insert(len(esv_table.columns), 'Seq', seq_col)

    ## save the final OTU table if selected
    if to_excel:
        wb = openpyxl.Workbook(write_only = True)
        ws = wb.create_sheet('ESV table')

        ## save the output line by line for optimized memory usage
        for row in tqdm(dataframe_to_rows(esv_table, index = False, header = True),
                                          total = len(esv_table.index),
                                          desc = '{}: Lines written to ESV table'.format(datetime.datetime.now().strftime("%H:%M:%S")),
                                          unit = ' lines'):
            ws.append(row)

        ## save the output (otu table)
        print('{}: Saving the ESV table to excel. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))
        wb.save(Path(project).joinpath('8_denoising', '{}_ESV_table.xlsx'.format(Path(project).stem)))
        wb.close()
        print('{}: ESV table saved to {}.'.format(datetime.datetime.now().strftime("%H:%M:%S"), Path(project).joinpath('8_denoising', '{}_ESV_table.xlsx'.format(Path(project).stem))))


    print('{}: Saving the ESV table to parquet. This may take a while.'.format(datetime.datetime.now().strftime("%H:%M:%S")))
    esv_table.to_parquet(Path(project).joinpath('8_denoising', '{}_ESV_table.parquet.snappy'.format(Path(project).stem)), index = False)
    print('{}: ESV table saved to {}.'.format(datetime.datetime.now().strftime("%H:%M:%S"), Path(project).joinpath('8_denoising', '{}_ESV_table.parquet.snappy'.format(Path(project).stem))))

    ## remove temporary files
    shutil.rmtree(Path(project).joinpath('8_denoising', 'temp'))

if __name__ == "__main__":
    main()
